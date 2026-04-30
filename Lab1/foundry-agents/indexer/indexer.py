import os
import io
import base64
import logging
import json
import re
import hashlib
import tempfile
from typing import List, Dict, Optional, Tuple

# detect-secrets for sensitive information scanning
try:
    from detect_secrets.core import scan
    from detect_secrets.settings import transient_settings
    DETECT_SECRETS_AVAILABLE = True
except ImportError:
    DETECT_SECRETS_AVAILABLE = False
    logger_fallback = logging.getLogger(__name__)
    logger_fallback.warning("detect-secrets library not available - sensitive info scanning will be disabled")
   
from azure.core.credentials import AzureKeyCredential
from azure.search.documents import SearchClient
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SimpleField,
    SearchFieldDataType,
    SearchableField,
    CorsOptions,
    SemanticConfiguration,
    SemanticPrioritizedFields,
    SemanticField,
    SemanticSearch,
    # Enhanced vector search imports for semantic hybrid search
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchProfile,
    SearchField,
    AzureOpenAIVectorizer,
    AzureOpenAIVectorizerParameters,
    SearchIndexerDataNoneIdentity,  # For system-assigned managed identity
)
from azure.storage.blob import BlobServiceClient
from azure.identity import DefaultAzureCredential, ManagedIdentityCredential

# Configure Azure SDK logging to reduce verbosity
# Suppress detailed HTTP request/response logging from Azure SDK
logging.getLogger('azure.core.pipeline.policies.http_logging_policy').setLevel(logging.WARNING)
logging.getLogger('azure.identity').setLevel(logging.WARNING)
logging.getLogger('azure.storage').setLevel(logging.WARNING)
logging.getLogger('azure.search').setLevel(logging.WARNING)

# Import logging and tracing configuration
try:
    from logging_config import get_logger
    from tracing_config import (
        get_tracer, add_span_attributes, record_search_operation, 
        initialize_tracing, trace_function
    )
    from opentelemetry.trace import Status, StatusCode
    
    # Setup logging first
    logger = get_logger(__name__)
    
    # Initialize tracing for the indexer module
    initialize_tracing()
    tracer = get_tracer()
    
    logger.info("Tracing initialized successfully for indexer module")
except Exception as ex:
    # Fallback logging setup
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(name)s: %(message)s',
        handlers=[logging.StreamHandler()]
    )
    logger = logging.getLogger(__name__)
    logger.warning(f"Failed to initialize tracing in indexer: {ex}")
    tracer = None
    # Create a dummy trace_function decorator when tracing fails
    def trace_function(name):
        def decorator(func):
            return func
        return decorator

# OpenAI SDK for embeddings
try:
    from openai import AzureOpenAI
except ImportError:
    AzureOpenAI = None

# -----------------------------
# Environment configuration
# -----------------------------
SEARCH_ENDPOINT = os.environ.get("AZURE_SEARCH_ENDPOINT", "")
SEARCH_INDEX = os.environ.get("AZURE_SEARCH_INDEX", "")  # fallback / legacy
SEARCH_KEY = os.environ.get("AZURE_SEARCH_ADMIN_KEY", "")
SEM_CONFIG_NAME = os.environ.get("AZURE_SEARCH_SEMANTIC_CONFIG")

# Storage: use account URL with Azure credentials or account name for constructing URL
AZ_STORAGE_ACCOUNT_URL = os.getenv("AZURE_STORAGE_ACCOUNT_URL")
# Add storage account name for constructing URL if needed
AZ_STORAGE_ACCOUNT_NAME = os.getenv("AZURE_STORAGE_ACCOUNT_NAME")

# Azure OpenAI (embeddings) configuration
# AOAI_ENDPOINT: Full endpoint URL for AzureOpenAI client (e.g., https://xxx.cognitiveservices.azure.com/openai/deployments/text-embedding-3-large/embeddings?api-version=2023-05-15)
# AOAI_ENDPOINT2: Base endpoint URL for AzureOpenAIVectorizer (e.g., https://xxx.openai.azure.com/)
AOAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "")
AOAI_ENDPOINT2 = os.getenv("AZURE_OPENAI_ENDPOINT2", "") 
AOAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-10-01-preview")
AOAI_EMBED_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", "")
# default to 3072 for text-embedding-3-large
AOAI_EMBED_DIM = int(os.getenv("AZURE_OPENAI_EMBED_DIM", "3072"))

# Managed Identity configuration
USE_MANAGED_IDENTITY_FOR_AOAI = os.getenv("USE_MANAGED_IDENTITY_FOR_AOAI", "true").lower() == "true"

# Sensitive information scanning configuration
ENABLE_SENSITIVE_SCAN = os.getenv("ENABLE_SENSITIVE_SCAN", "true").lower() == "true"
SKIP_SENSITIVE_DOCUMENTS = os.getenv("SKIP_SENSITIVE_DOCUMENTS", "true").lower() == "true"

logger.info("=" * 80)
logger.info("INDEXER CONFIGURATION")
logger.info("=" * 80)
logger.info(f"Azure Search Endpoint: {SEARCH_ENDPOINT}")
logger.info(f"Azure Search Index: {SEARCH_INDEX}")
logger.info(f"Using Managed Identity for Search: {not SEARCH_KEY or os.getenv('USE_MANAGED_IDENTITY', 'true').lower() == 'true'}")
logger.info(f"Azure OpenAI Endpoint: {AOAI_ENDPOINT}")
logger.info(f"Azure OpenAI Endpoint2: {AOAI_ENDPOINT2}")
logger.info(f"Azure OpenAI Deployment: {AOAI_EMBED_DEPLOYMENT}")
logger.info(f"Azure OpenAI API Version: {AOAI_API_VERSION}")
logger.info(f"Embedding Dimensions: {AOAI_EMBED_DIM}")
logger.info(f"Using Managed Identity for Azure OpenAI: {USE_MANAGED_IDENTITY_FOR_AOAI}")
logger.info(f"Semantic Config Name: {SEM_CONFIG_NAME}")
logger.info(f"Storage Account URL: {AZ_STORAGE_ACCOUNT_URL}")
logger.info(f"Storage Account Name: {AZ_STORAGE_ACCOUNT_NAME}")
logger.info(f"Sensitive Info Scanning Enabled: {ENABLE_SENSITIVE_SCAN}")
logger.info(f"Skip Sensitive Documents: {SKIP_SENSITIVE_DOCUMENTS}")
#logger.info(f"detect-secrets Available: {DETECT_SECRETS_AVAILABLE}")
logger.info("detect-secrets integration configured")
logger.info("=" * 80)


# -----------------------------
# Sensitive Information Detection using detect-secrets
# -----------------------------

def _scan_for_sensitive_info(text: str, source_name: str = "document") -> Tuple[bool, List[Dict]]:
    """
    Scan text for sensitive information (keys, passwords, secrets) using detect-secrets.
    
    Args:
        text: The text content to scan
        source_name: Name of the source document for logging purposes
        
    Returns:
        Tuple of (has_secrets: bool, secrets_found: List[Dict])
        - has_secrets: True if sensitive information was detected
        - secrets_found: List of dictionaries containing secret details (type, line number)
    """
    if not ENABLE_SENSITIVE_SCAN:
        logger.debug(f"Sensitive info scanning disabled - skipping scan for {source_name}")
        return False, []
    
    if not DETECT_SECRETS_AVAILABLE:
        logger.warning("detect-secrets library not available - skipping sensitive info scan")
        return False, []
    
    secrets_found = []
    
    try:
        # Create a temporary file to scan (detect-secrets works with files)
        with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as tmp_file:
            tmp_file.write(text)
            tmp_file_path = tmp_file.name
        
        try:
            # Use transient_settings to configure the scan with all default plugins
            with transient_settings({
                'plugins_used': [
                    {'name': 'ArtifactoryDetector'},
                    {'name': 'AWSKeyDetector'},
                    {'name': 'AzureStorageKeyDetector'},
                    {'name': 'BasicAuthDetector'},
                    {'name': 'CloudantDetector'},
                    {'name': 'DiscordBotTokenDetector'},
                    {'name': 'GitHubTokenDetector'},
                    {'name': 'Base64HighEntropyString', 'limit': 4.5},
                    {'name': 'HexHighEntropyString', 'limit': 3.0},
                    {'name': 'IbmCloudIamDetector'},
                    {'name': 'IbmCosHmacDetector'},
                    {'name': 'JwtTokenDetector'},
                    {'name': 'KeywordDetector'},
                    {'name': 'MailchimpDetector'},
                    {'name': 'NpmDetector'},
                    {'name': 'OpenAIDetector'},
                    {'name': 'PrivateKeyDetector'},
                    {'name': 'SendGridDetector'},
                    {'name': 'SlackDetector'},
                    {'name': 'StripeDetector'},
                    {'name': 'TwilioKeyDetector'},
                ]
            }):
                # Scan the file for secrets
                secrets = scan.scan_file(tmp_file_path)
                
                # Process found secrets
                for secret in secrets:
                    secret_info = {
                        'type': secret.type,
                        'line_number': secret.line_number,
                        'source': source_name
                    }
                    secrets_found.append(secret_info)
                    logger.debug(f"  SENSITIVE INFO DETECTED in {source_name}: {secret.type} at line {secret.line_number}")
                    
        finally:
            # Clean up temporary file
            try:
                os.unlink(tmp_file_path)
            except Exception:
                pass
                
    except Exception as ex:
        logger.error(f"Error scanning {source_name} for sensitive info: {ex}", exc_info=True)
        # Return False to allow indexing to continue if scanning fails
        return False, []
    
    has_secrets = len(secrets_found) > 0
    
    if has_secrets:
        # Console: generic warning without details
        logger.warning(f"  Found {len(secrets_found)} potential secret(s) in {source_name}")
        for secret in secrets_found:
            logger.debug(f"    - {secret['type']} at line {secret['line_number']}")
    else:
        logger.debug(f"  No sensitive information detected in {source_name}")
    
    return has_secrets, secrets_found


def _scan_for_cwe_vulnerabilities(text: str, source_name: str = "document") -> Tuple[bool, List[Dict]]:
    """
    Scan text for CWE vulnerabilities in Kubernetes/container configurations.
    
    Detects:
    - CWE-653: Improper Isolation (hostPID, hostIPC, hostNetwork)
    - CWE-284: Improper Access Control (privileged, allowPrivilegeEscalation)
    - Missing resource limits (requests/limits for CPU/memory)
    
    Args:
        text: The text content to scan
        source_name: Name of the source document for logging
        
    Returns:
        Tuple of (has_vulnerabilities: bool, vulnerabilities_found: List[Dict])
    """
    if not ENABLE_SENSITIVE_SCAN:
        return False, []
    
    vulnerabilities = []
    text_lower = text.lower()
    
    # CWE-653: Improper Isolation or Compartmentalization
    cwe_653_patterns = [
        (r'hostpid\s*:\s*true', 'CWE-653: hostPID enabled - allows access to host processes'),
        (r'hostipc\s*:\s*true', 'CWE-653: hostIPC enabled - allows access to host IPC'),
        (r'hostnetwork\s*:\s*true', 'CWE-653: hostNetwork enabled - uses host network namespace'),
    ]
    
    # CWE-284: Improper Access Control
    cwe_284_patterns = [
        (r'privileged\s*:\s*true', 'CWE-284: Privileged container - full access to host resources'),
        (r'allowprivilegeescalation\s*:\s*true', 'CWE-284: Privilege escalation allowed - setuid binary exploitation risk'),
        (r'runasuser\s*:\s*0', 'CWE-284: Running as root user (UID 0)'),
    ]
    
    # Check for missing resource limits
    has_container_spec = 'container' in text_lower and ('image:' in text_lower or 'name:' in text_lower)
    has_resources = 'resources:' in text_lower
    has_limits = 'limits:' in text_lower
    has_requests = 'requests:' in text_lower
    
    # Scan for CWE-653 patterns
    for pattern, description in cwe_653_patterns:
        matches = re.finditer(pattern, text_lower, re.IGNORECASE)
        for match in matches:
            line_num = text[:match.start()].count('\n') + 1
            vulnerabilities.append({
                'type': 'CWE-653',
                'description': description,
                'line_number': line_num,
                'source': source_name,
                'pattern': pattern
            })
            logger.debug(f"  CWE VULNERABILITY DETECTED in {source_name}: {description} at line {line_num}")
    
    # Scan for CWE-284 patterns
    for pattern, description in cwe_284_patterns:
        matches = re.finditer(pattern, text_lower, re.IGNORECASE)
        for match in matches:
            line_num = text[:match.start()].count('\n') + 1
            vulnerabilities.append({
                'type': 'CWE-284',
                'description': description,
                'line_number': line_num,
                'source': source_name,
                'pattern': pattern
            })
            logger.debug(f"  CWE VULNERABILITY DETECTED in {source_name}: {description} at line {line_num}")
    
    # Check for missing resource limits (common security issue)
    if has_container_spec and not has_resources:
        vulnerabilities.append({
            'type': 'RESOURCE-LIMIT',
            'description': 'Container lacks resource limits (CPU/memory) - risk of resource exhaustion',
            'line_number': 0,
            'source': source_name,
            'pattern': 'missing resources'
        })
        logger.debug(f"  CWE VULNERABILITY DETECTED in {source_name}: Missing resource limits")
    elif has_container_spec and has_resources:
        if not has_limits:
            vulnerabilities.append({
                'type': 'RESOURCE-LIMIT',
                'description': 'Container missing resource limits - risk of unbounded resource usage',
                'line_number': 0,
                'source': source_name,
                'pattern': 'missing limits'
            })
            logger.debug(f"  CWE VULNERABILITY DETECTED in {source_name}: Missing resource limits")
        if not has_requests:
            vulnerabilities.append({
                'type': 'RESOURCE-LIMIT',
                'description': 'Container missing resource requests - improper scheduling',
                'line_number': 0,
                'source': source_name,
                'pattern': 'missing requests'
            })
            logger.debug(f"  CWE VULNERABILITY DETECTED in {source_name}: Missing resource requests")
    
    has_vulnerabilities = len(vulnerabilities) > 0
    
    if has_vulnerabilities:
        # Console: generic warning without details
        logger.warning(f"  Found {len(vulnerabilities)} CWE vulnerability/vulnerabilities in {source_name}")
        for vuln in vulnerabilities:
            logger.debug(f"    - {vuln['type']}: {vuln['description']}")
    else:
        logger.debug(f"  No CWE vulnerabilities detected in {source_name}")
    
    return has_vulnerabilities, vulnerabilities


def _should_skip_document(text: str, source_name: str = "document") -> Tuple[bool, List[Dict]]:
    """
    Check if a document should be skipped due to sensitive information or security vulnerabilities.
    
    Args:
        text: The text content to check
        source_name: Name of the source document for logging
        
    Returns:
        Tuple of (should_skip: bool, issues_found: List[Dict])
        issues_found contains both secrets and vulnerabilities
    """
    # Scan for sensitive information (API keys, passwords, etc.)
    has_secrets, secrets_found = _scan_for_sensitive_info(text, source_name)
    
    # Scan for CWE vulnerabilities (Kubernetes/container security issues)
    has_vulnerabilities, vulnerabilities_found = _scan_for_cwe_vulnerabilities(text, source_name)
    
    # Combine all issues
    all_issues = secrets_found + vulnerabilities_found
    has_issues = has_secrets or has_vulnerabilities
    
    if has_issues and SKIP_SENSITIVE_DOCUMENTS:
        issue_summary = []
        if has_secrets:
            issue_summary.append(f"{len(secrets_found)} sensitive item(s)")
        if has_vulnerabilities:
            issue_summary.append(f"{len(vulnerabilities_found)} CWE vulnerability/vulnerabilities")
        
        logger.warning(f"  SKIPPING {source_name}: Contains {', '.join(issue_summary)}")
        return True, all_issues
    elif has_issues:
        issue_summary = []
        if has_secrets:
            issue_summary.append(f"{len(secrets_found)} sensitive item(s)")
        if has_vulnerabilities:
            issue_summary.append(f"{len(vulnerabilities_found)} CWE vulnerability/vulnerabilities")
        
        logger.warning(f"  FLAGGING {source_name}: Contains {', '.join(issue_summary)} (not skipping - SKIP_SENSITIVE_DOCUMENTS=false)")
        return False, all_issues
    
    return False, []


# -----------------------------
# Enhanced credential management for managed identity
# -----------------------------

def _get_search_credential():
    """
    Get search credential using managed identity first, then fall back to API key.
    
    Returns:
        Credential object for Azure Search authentication
    """
    # Try managed identity first (recommended for production) but fall back on permission errors
    use_managed_identity = os.getenv("USE_MANAGED_IDENTITY", "true").lower() == "true"
    
    if SEARCH_KEY and not use_managed_identity:
        logger.info("Using API key for Azure Search authentication (explicitly configured)")
        return AzureKeyCredential(SEARCH_KEY)
    elif SEARCH_KEY:
        logger.info("Using API key for Azure Search authentication (fallback from managed identity)")
        return AzureKeyCredential(SEARCH_KEY)
    else:
        logger.info("Using managed identity for Azure Search authentication")
        return DefaultAzureCredential(exclude_shared_token_cache_credential=True)


def _get_search_client(index_name: str) -> SearchClient:
    """
    Create a SearchClient with appropriate authentication.
    
    Args:
        index_name: Name of the search index
        
    Returns:
        SearchClient instance
    """
    credential = _get_search_credential()
    return SearchClient(SEARCH_ENDPOINT, index_name, credential)


def _get_search_index_client() -> SearchIndexClient:
    """
    Create a SearchIndexClient with appropriate authentication.
    
    Returns:
        SearchIndexClient instance
    """
    credential = _get_search_credential()
    return SearchIndexClient(SEARCH_ENDPOINT, credential)


# -----------------------------
# Helpers matching existing repo with enhanced tracing
# -----------------------------

@trace_function("blob_client_creation")
def _blob_client() -> BlobServiceClient:
    """
    Create BlobServiceClient with the following priority:
    1. Account URL with Managed Identity (recommended)
    2. Account name to construct URL with Managed Identity
    """
    # Use ManagedIdentityCredential in Azure, fall back to DefaultAzureCredential locally
    try:
        credential = ManagedIdentityCredential()
        credential.get_token("https://storage.azure.com/.default")
    except Exception:
        credential = DefaultAzureCredential(exclude_shared_token_cache_credential=True)
    
    # Option 1: Use account URL with Managed Identity or SAS
    if AZ_STORAGE_ACCOUNT_URL:
        # If SAS token is in URL (contains '?'), use it directly
        if "?" in AZ_STORAGE_ACCOUNT_URL:
            return BlobServiceClient(account_url=AZ_STORAGE_ACCOUNT_URL)
        # Otherwise use Managed Identity
        return BlobServiceClient(
            account_url=AZ_STORAGE_ACCOUNT_URL, 
            credential=credential
        )
    
    # Option 2: Construct URL from account name and use Managed Identity
    if AZ_STORAGE_ACCOUNT_NAME:
        account_url = f"https://{AZ_STORAGE_ACCOUNT_NAME}.blob.core.windows.net"
        return BlobServiceClient(
            account_url=account_url,
            credential=credential
        )
    
    # If none are set, raise an error
    raise RuntimeError(
        "Storage configuration missing. Set one of: "
        "AZURE_STORAGE_ACCOUNT_URL or AZURE_STORAGE_ACCOUNT_NAME"
    )


@trace_function("get_blob_metadata")
def _get_blob_metadata(container: str, blob_name: str) -> Dict[str, str]:
    """
    Retrieve the metadata key-value pairs from a blob in Azure Storage.
    
    Args:
        container: The blob container name
        blob_name: The blob name
        
    Returns:
        Dictionary containing the blob's metadata key-value pairs.
        Returns empty dict if no metadata is found or if an error occurs.
    """
    try:
        bc = _blob_client().get_blob_client(container=container, blob=blob_name)
        properties = bc.get_blob_properties()
        metadata = properties.metadata or {}
        
        if metadata:
            logger.debug(f"Retrieved blob metadata for {blob_name}: {metadata}")
        else:
            logger.debug(f"No metadata found for blob: {blob_name}")
            
        return metadata
    except Exception as ex:
        logger.warning(f"Failed to retrieve metadata for blob {container}/{blob_name}: {ex}")
        return {}


def _format_blob_metadata_as_string(metadata: Dict[str, str]) -> str:
    """
    Format blob metadata dictionary as a searchable/filterable string.
    
    Args:
        metadata: Dictionary of blob metadata key-value pairs
        
    Returns:
        JSON-formatted string representation of the metadata for indexing.
        Returns empty string if no metadata.
    """
    if not metadata:
        return ""
    
    try:
        # Return as JSON string for easy parsing and filtering
        return json.dumps(metadata, ensure_ascii=False)
    except Exception as ex:
        logger.warning(f"Failed to format metadata as string: {ex}")
        # Fallback: simple key=value format
        return ", ".join([f"{k}={v}" for k, v in metadata.items()])


@trace_function("aoai_client_creation") 
def _get_aoai_client():
    """Create Azure OpenAI client using OpenAI SDK (AzureOpenAI). 
    
    Uses managed identity authentication (ManagedIdentityCredential).
    Returns None if not configured.
    """
    logger.info("Creating Azure OpenAI client...")
    
    if not AOAI_ENDPOINT or not AOAI_EMBED_DEPLOYMENT:
        logger.warning("Azure OpenAI embeddings not configured - missing AOAI_ENDPOINT or AOAI_EMBED_DEPLOYMENT")
        logger.warning("Vector search will be disabled")
        return None
        
    try:
        if AzureOpenAI is None:
            logger.error("OpenAI SDK not available - install with: pip install openai")
            return None
        
        logger.info("Creating Azure OpenAI client with Managed Identity authentication")
        from azure.identity import ManagedIdentityCredential, get_bearer_token_provider
        
        # Create managed identity credential
        credential = ManagedIdentityCredential()
        
        # Create token provider that automatically refreshes tokens
        token_provider = get_bearer_token_provider(
            credential,
            "https://cognitiveservices.azure.com/.default"
        )
        
        logger.info("Successfully created token provider for Azure OpenAI via Managed Identity")
        
        # Create client with token provider (automatically refreshes tokens)
        client = AzureOpenAI(
            azure_ad_token_provider=token_provider,
            api_version=AOAI_API_VERSION,
            azure_endpoint=AOAI_ENDPOINT
        )
        logger.info(f"Azure OpenAI client created successfully with Managed Identity")
        logger.info(f"  Endpoint: {AOAI_ENDPOINT}")
        logger.info(f"  Deployment: {AOAI_EMBED_DEPLOYMENT}")
        logger.info(f"  API Version: {AOAI_API_VERSION}")
        logger.info(f"  Token will be refreshed automatically")
        
        return client
            
    except Exception as e:
        logger.error(f"Failed to create Azure OpenAI client: {e}", exc_info=True)
        return None


@trace_function("embed_texts")
def _embed_texts(texts: List[str]) -> List[List[float] | None]:
    """Return embeddings for texts via AzureOpenAI. If not available, returns [None,...]."""
    logger.info(f"Starting embedding generation for {len(texts)} text chunks...")
    
    if tracer:
        with tracer.start_as_current_span("embed_texts_execution") as span:
            try:
                add_span_attributes(span, {
                    "embedding.text_count": len(texts),
                    "embedding.deployment": AOAI_EMBED_DEPLOYMENT,
                    "embedding.dimensions": AOAI_EMBED_DIM
                })

                client = _get_aoai_client()
                if not client:
                    logger.warning("Azure OpenAI client not available - skipping embedding generation")
                    span.add_event("aoai_client_unavailable", {"reason": "No client configured"})
                    return [None] * len(texts)

                vectors: List[List[float] | None] = []
                batch_size = 16
                total_batches = (len(texts) + batch_size - 1) // batch_size
                
                logger.info(f"Processing {total_batches} batches of embeddings (batch size: {batch_size})...")
                
                for batch_idx, i in enumerate(range(0, len(texts), batch_size)):
                    with tracer.start_as_current_span(f"embed_batch_{batch_idx + 1}") as batch_span:
                        try:
                            batch = texts[i : i + batch_size]
                            
                            logger.debug(f"  Batch {batch_idx + 1}/{total_batches}: Processing {len(batch)} texts...")
                            
                            add_span_attributes(batch_span, {
                                "embedding.batch_index": batch_idx + 1,
                                "embedding.batch_size": len(batch),
                                "embedding.total_batches": total_batches
                            })

                            resp = client.embeddings.create(model=AOAI_EMBED_DEPLOYMENT, input=batch)
                            vecs = [d.embedding for d in resp.data]
                            
                            for v in vecs:
                                if isinstance(v, list) and len(v) != AOAI_EMBED_DIM:
                                    logger.warning(f"Embedding dimension {len(v)} != expected {AOAI_EMBED_DIM}")
                            
                            vectors.extend(vecs)
                            batch_span.set_status(Status(StatusCode.OK))
                            logger.debug(f"  Batch {batch_idx + 1}/{total_batches}: Successfully generated {len(vecs)} embeddings")
                            
                        except Exception as e:  # pragma: no cover
                            batch_span.record_exception(e)
                            batch_span.set_status(Status(StatusCode.ERROR, str(e)))
                            logger.error(f"  Batch {batch_idx + 1}/{total_batches}: Embedding request failed: {e}", exc_info=True)
                            vectors.extend([None] * len(batch))

                successful = sum(1 for v in vectors if v is not None)
                failed = sum(1 for v in vectors if v is None)
                
                logger.info(f"Embedding generation complete: {successful} successful, {failed} failed")
                
                add_span_attributes(span, {
                    "embedding.successful_vectors": successful,
                    "embedding.failed_vectors": failed
                })

                span.set_status(Status(StatusCode.OK))
                return vectors
                
            except Exception as ex:
                span.record_exception(ex)
                span.set_status(Status(StatusCode.ERROR, str(ex)))
                logger.error(f"Embedding generation failed: {ex}", exc_info=True)
                raise
    else:
        # Fallback without tracing
        logger.info("Processing embeddings without tracing...")
        
        client = _get_aoai_client()
        if not client:
            logger.warning("Azure OpenAI client not available - skipping embedding generation")
            return [None] * len(texts)

        vectors: List[List[float] | None] = []
        batch_size = 16
        total_batches = (len(texts) + batch_size - 1) // batch_size
        
        logger.info(f"Processing {total_batches} batches of embeddings (batch size: {batch_size})...")
        
        for batch_idx, i in enumerate(range(0, len(texts), batch_size)):
            batch = texts[i : i + batch_size]
            try:
                logger.debug(f"  Batch {batch_idx + 1}/{total_batches}: Processing {len(batch)} texts...")
                resp = client.embeddings.create(model=AOAI_EMBED_DEPLOYMENT, input=batch)
                vecs = [d.embedding for d in resp.data]
                for v in vecs:
                    if isinstance(v, list) and len(v) != AOAI_EMBED_DIM:
                        logger.warning(f"Embedding dimension {len(v)} != expected {AOAI_EMBED_DIM}")
                vectors.extend(vecs)
                logger.debug(f"  Batch {batch_idx + 1}/{total_batches}: Successfully generated {len(vecs)} embeddings")
            except Exception as e:  # pragma: no cover
                logger.error(f"  Batch {batch_idx + 1}/{total_batches}: Embedding request failed: {e}", exc_info=True)
                vectors.extend([None] * len(batch))
        
        successful = sum(1 for v in vectors if v is not None)
        failed = sum(1 for v in vectors if v is None)
        logger.info(f"Embedding generation complete: {successful} successful, {failed} failed")
        
        return vectors


def _sanitize_index_name(app_id: str) -> str:
    """Convert arbitrary appId to a valid Azure AI Search index name.

    Rules: lowercase, alphanumerics or dashes; must start/end with alphanumeric; length 2-128.
    """
    import re, hashlib
    base = app_id.lower().strip()
    base = re.sub(r"[^a-z0-9-]", "-", base)            # invalid chars -> dash
    base = re.sub(r"-+", "-", base)                     # collapse dashes
    base = base.strip("-")                                 # trim
    if not base or not base[0].isalnum():
        base = f"app-{hashlib.sha1(app_id.encode()).hexdigest()[:8]}"
    if len(base) < 2:
        base = f"{base}ix"
    if len(base) > 128:
        base = base[:128]
    return base


@trace_function("create_or_update_index")
def create_or_update_index(app_id: str, force_recreate: bool = True) -> str:
    """Create (or recreate) a search index whose name is derived from app_id.
    
    Enhanced to support semantic hybrid search (semantic + vector) following Microsoft's best practices.
    If force_recreate=True and the index exists, delete it first so schema changes are applied.
    Returns the index name actually used.
    """
    if tracer:
        with tracer.start_as_current_span("create_update_index_execution") as span:
            try:
                if not SEARCH_ENDPOINT:
                    error_msg = "AZURE_SEARCH_ENDPOINT must be set"
                    span.set_status(Status(StatusCode.ERROR, error_msg))
                    raise RuntimeError(error_msg)

                index_name = _sanitize_index_name(app_id)
                
                add_span_attributes(span, {
                    "search.app_id": app_id,
                    "search.index_name": index_name,
                    "search.force_recreate": force_recreate,
                    "search.endpoint": SEARCH_ENDPOINT,
                    "search.using_managed_identity": not SEARCH_KEY or os.getenv("USE_MANAGED_IDENTITY", "true").lower() == "true"
                })

                # Use enhanced credential management
                idx_client = _get_search_index_client()

                # Enhanced field definition following Microsoft's approach
                fields = [
                    SimpleField(name="id", type=SearchFieldDataType.String, key=True),
                    SimpleField(name="appId", type=SearchFieldDataType.String, filterable=True, sortable=True, facetable=True),
                    SearchableField(name="title", type=SearchFieldDataType.String, analyzer_name="en.lucene"),
                    SearchableField(name="content", type=SearchFieldDataType.String, analyzer_name="en.lucene"),
                    SimpleField(name="source", type=SearchFieldDataType.String, filterable=True),
                    SearchableField(name="path", type=SearchFieldDataType.String, filterable=True),  # Standard analyzer for path matching
                    SimpleField(name="chunkId", type=SearchFieldDataType.String, filterable=True),
                    # Additional fields for better semantic search
                    SearchableField(name="filepath", type=SearchFieldDataType.String),
                    SearchableField(name="url", type=SearchFieldDataType.String),
                    SearchableField(name="metadata", type=SearchFieldDataType.String, filterable=True)
                ]

                # Enhanced vector field configuration
                vector_search = None
                vectorizers = []
                has_vector_field = False
                
                with tracer.start_as_current_span("configure_vector_search") as vector_span:
                    try:
                        if AOAI_EMBED_DIM > 0 and AOAI_ENDPOINT and AOAI_EMBED_DEPLOYMENT:
                            has_vector_field = True
                            
                            logger.info("Configuring vector search with Azure OpenAI vectorizer...")
                            logger.info(f"  Embedding dimensions: {AOAI_EMBED_DIM}")
                            logger.info(f"  Deployment: {AOAI_EMBED_DEPLOYMENT}")
                            logger.info(f"  Endpoint: {AOAI_ENDPOINT2 or AOAI_ENDPOINT}")
                            logger.info(f"  Using Managed Identity: {USE_MANAGED_IDENTITY_FOR_AOAI}")
                            
                            # Create Azure OpenAI vectorizer for integrated vector search
                            # This is required for VECTOR_SEMANTIC_HYBRID query type in agents
                            # Uses Search service's system-assigned managed identity
                            logger.info("  Vectorizer will use Search service's System Managed Identity")
                            vectorizer_params = AzureOpenAIVectorizerParameters(
                                resource_url=AOAI_ENDPOINT2 or AOAI_ENDPOINT,
                                deployment_name=AOAI_EMBED_DEPLOYMENT,
                                model_name=AOAI_EMBED_DEPLOYMENT,
                                auth_identity=SearchIndexerDataNoneIdentity()  # Use search service's system MI
                            )
                            
                            vectorizer = AzureOpenAIVectorizer(
                                vectorizer_name="myOpenAI",
                                parameters=vectorizer_params
                            )
                            vectorizers.append(vectorizer)
                            
                            logger.info("Vectorizer configured successfully")
                            
                            # Add vector field with proper configuration and integrated vectorizer
                            fields.append(
                                SearchField(
                                    name="contentVector",
                                    type=SearchFieldDataType.Collection(SearchFieldDataType.Single),
                                    searchable=True,
                                    hidden=False,
                                    vector_search_dimensions=AOAI_EMBED_DIM,
                                    vector_search_profile_name="myHnswProfile",
                                )
                            )
                            
                            logger.info("Vector field 'contentVector' added to index schema")
                            
                            # Enhanced vector search configuration following Microsoft's pattern
                            vector_search = VectorSearch(
                                algorithms=[
                                    HnswAlgorithmConfiguration(
                                        name="myHnsw",
                                        parameters={
                                            "m": 4,
                                            "efConstruction": 400,
                                            "efSearch": 500,
                                            "metric": "cosine"
                                        }
                                    )
                                ],
                                profiles=[
                                    VectorSearchProfile(
                                        name="myHnswProfile",
                                        algorithm_configuration_name="myHnsw",
                                        vectorizer_name="myOpenAI"  # Link to the integrated vectorizer
                                    )
                                ],
                                vectorizers=vectorizers  # Add the vectorizers to enable integrated search
                            )
                            
                            logger.info("Vector search configuration created with HNSW algorithm")
                            logger.info("  Algorithm: myHnsw (m=4, efConstruction=400, efSearch=500, metric=cosine)")
                            logger.info("  Profile: myHnswProfile")
                            logger.info("  Integrated vectorizer: myOpenAI")

                        add_span_attributes(vector_span, {
                            "vector.enabled": has_vector_field,
                            "vector.dimensions": AOAI_EMBED_DIM if has_vector_field else None,
                            "vector.deployment": AOAI_EMBED_DEPLOYMENT if has_vector_field else None
                        })

                        vector_span.set_status(Status(StatusCode.OK))
                        
                    except Exception as ex:
                        vector_span.record_exception(ex)
                        vector_span.set_status(Status(StatusCode.ERROR, str(ex)))
                        raise

                # Enhanced semantic search configuration
                semantic_search = None
                semantic_config_name = SEM_CONFIG_NAME or "my-semantic-config"
                
                # Always create semantic configuration for better search quality
                semantic_search = SemanticSearch(
                    configurations=[
                        SemanticConfiguration(
                            name=semantic_config_name,
                            prioritized_fields=SemanticPrioritizedFields(
                                title_field=SemanticField(field_name="title"),
                                content_fields=[
                                    SemanticField(field_name="content"),
                                    SemanticField(field_name="metadata")
                                ],
                                keywords_fields=[
                                    SemanticField(field_name="chunkId"),
                                    SemanticField(field_name="source")
                                ]
                            )
                        )
                    ]
                )

                # Create index with enhanced configuration
                index = SearchIndex(
                    name=index_name,
                    fields=fields,
                    cors_options=CorsOptions(allowed_origins=["*"], max_age_in_seconds=60),
                    vector_search=vector_search,
                    semantic_search=semantic_search
                )

                # Recreate if required
                with tracer.start_as_current_span("index_operations") as index_ops_span:
                    try:
                        existing = idx_client.get_index(index_name)
                        if existing and force_recreate:
                            logging.info("Deleting existing index '%s' to recreate", index_name)
                            idx_client.delete_index(index_name)
                            index_ops_span.add_event("index_deleted", {"index_name": index_name})
                    except Exception:
                        # Not found -> safe to create
                        index_ops_span.add_event("index_not_found", {"index_name": index_name})

                    # Create (again) - if it still exists and force_recreate False, this will raise; catch & ignore
                    try:
                        result = idx_client.create_index(index)
                        logging.info(f"Created search index '{result.name}' with semantic hybrid search capabilities")
                        
                        # Log configuration details
                        config_details = {
                            "vector_enabled": has_vector_field,
                            "semantic_enabled": True,
                            "vector_dimensions": AOAI_EMBED_DIM if has_vector_field else None,
                            "semantic_config": semantic_config_name,
                            "supports_hybrid": has_vector_field,  # Vector + Semantic = Hybrid
                            "integrated_vectorizer": bool(vectorizers),  # Required for agent VECTOR_SEMANTIC_HYBRID
                            "vectorizer_deployment": AOAI_EMBED_DEPLOYMENT if has_vector_field else None
                        }
                        logging.info(f"Index configuration: {config_details}")
                        
                        add_span_attributes(index_ops_span, {
                            "search.operation": "created",
                            **{f"search.config.{k}": str(v) for k, v in config_details.items()}
                        })
                        
                    except Exception as e:
                        if "already exists" in str(e).lower():
                            logging.info(f"Index {index_name} already exists")
                            index_ops_span.add_event("index_already_exists", {"index_name": index_name})
                        else:
                            logging.error(f"Failed to create index {index_name}: {e}")
                            index_ops_span.record_exception(e)
                            index_ops_span.set_status(Status(StatusCode.ERROR, str(e)))
                            raise
                    
                    index_ops_span.set_status(Status(StatusCode.OK))

                span.set_status(Status(StatusCode.OK))
                return index_name
                
            except Exception as ex:
                span.record_exception(ex)
                span.set_status(Status(StatusCode.ERROR, str(ex)))
                raise
    else:
        # Fallback without tracing - original implementation
        if not SEARCH_ENDPOINT:
            raise RuntimeError("AZURE_SEARCH_ENDPOINT must be set")

        index_name = _sanitize_index_name(app_id)
        idx_client = _get_search_index_client()

        # Enhanced field definition following Microsoft's approach
        fields = [
            SimpleField(name="id", type=SearchFieldDataType.String, key=True),
            SimpleField(name="appId", type=SearchFieldDataType.String, filterable=True, sortable=True, facetable=True),
            SearchableField(name="title", type=SearchFieldDataType.String, analyzer_name="en.lucene"),
            SearchableField(name="content", type=SearchFieldDataType.String, analyzer_name="en.lucene"),
            SimpleField(name="source", type=SearchFieldDataType.String, filterable=True),
            SearchableField(name="path", type=SearchFieldDataType.String, filterable=True),  # Standard analyzer for path matching
            SimpleField(name="chunkId", type=SearchFieldDataType.String, filterable=True),
            # Additional fields for better semantic search
            SearchableField(name="filepath", type=SearchFieldDataType.String),
            SearchableField(name="url", type=SearchFieldDataType.String),
            SearchableField(name="metadata", type=SearchFieldDataType.String, filterable=True),
        ]

        # Enhanced vector field configuration
        vector_search = None
        vectorizers = []
        has_vector_field = False
        
        if AOAI_EMBED_DIM > 0 and AOAI_ENDPOINT and AOAI_EMBED_DEPLOYMENT:
            has_vector_field = True
            
            logger.info("Configuring vector search with Azure OpenAI vectorizer (fallback path)...")
            logger.info(f"  Embedding dimensions: {AOAI_EMBED_DIM}")
            logger.info(f"  Deployment: {AOAI_EMBED_DEPLOYMENT}")
            logger.info(f"  Endpoint: {AOAI_ENDPOINT2 or AOAI_ENDPOINT}")
            logger.info(f"  Using Managed Identity: {USE_MANAGED_IDENTITY_FOR_AOAI}")
            
            # Create Azure OpenAI vectorizer for integrated vector search
            # This is required for VECTOR_SEMANTIC_HYBRID query type in agents
            # Uses Search service's system-assigned managed identity
            logger.info("  Vectorizer will use Search service's System Managed Identity")
            vectorizer_params = AzureOpenAIVectorizerParameters(
                resource_url=AOAI_ENDPOINT2 or AOAI_ENDPOINT,
                deployment_name=AOAI_EMBED_DEPLOYMENT,
                model_name=AOAI_EMBED_DEPLOYMENT,
                auth_identity=SearchIndexerDataNoneIdentity()  # Use search service's system MI
            )
            
            vectorizer = AzureOpenAIVectorizer(
                vectorizer_name="myOpenAI",
                parameters=vectorizer_params
            )
            vectorizers.append(vectorizer)
            
            logger.info("Vectorizer configured successfully")
            
            # Add vector field with proper configuration and integrated vectorizer
            fields.append(
                SearchField(
                    name="contentVector",
                    type=SearchFieldDataType.Collection(SearchFieldDataType.Single),
                    searchable=True,
                    hidden=False,
                    vector_search_dimensions=AOAI_EMBED_DIM,
                    vector_search_profile_name="myHnswProfile",
                )
            )
            
            # Enhanced vector search configuration following Microsoft's pattern
            vector_search = VectorSearch(
                algorithms=[
                    HnswAlgorithmConfiguration(
                        name="myHnsw",
                        parameters={
                            "m": 4,
                            "efConstruction": 400,
                            "efSearch": 500,
                            "metric": "cosine"
                        }
                    )
                ],
                profiles=[
                    VectorSearchProfile(
                        name="myHnswProfile",
                        algorithm_configuration_name="myHnsw",
                        vectorizer_name="myOpenAI"  # Link to the integrated vectorizer
                    )
                ],
                vectorizers=vectorizers  # Add the vectorizers to enable integrated search
            )

        # Enhanced semantic search configuration
        semantic_search = None
        semantic_config_name = SEM_CONFIG_NAME or "my-semantic-config"
        
        # Always create semantic configuration for better search quality
        semantic_search = SemanticSearch(
            configurations=[
                SemanticConfiguration(
                    name=semantic_config_name,
                    prioritized_fields=SemanticPrioritizedFields(
                        title_field=SemanticField(field_name="title"),
                        content_fields=[
                            SemanticField(field_name="content"),
                            SemanticField(field_name="metadata")
                        ],
                        keywords_fields=[
                            SemanticField(field_name="chunkId"),
                            SemanticField(field_name="source")
                        ]
                    )
                )
            ]
        )

        # Create index with enhanced configuration
        index = SearchIndex(
            name=index_name,
            fields=fields,
            cors_options=CorsOptions(allowed_origins=["*"], max_age_in_seconds=60),
            vector_search=vector_search,
            semantic_search=semantic_search
        )

        # Recreate if required
        try:
            existing = idx_client.get_index(index_name)
            if existing and force_recreate:
                logging.info("Deleting existing index '%s' to recreate", index_name)
                idx_client.delete_index(index_name)
        except Exception:
            # Not found -> safe to create
            pass

        # Create (again) - if it still exists and force_recreate False, this will raise; catch & ignore
        try:
            result = idx_client.create_index(index)
            logging.info(f"Created search index '{result.name}' with semantic hybrid search capabilities")
            
            # Log configuration details
            config_details = {
                "vector_enabled": has_vector_field,
                "semantic_enabled": True,
                "vector_dimensions": AOAI_EMBED_DIM if has_vector_field else None,
                "semantic_config": semantic_config_name,
                "supports_hybrid": has_vector_field,  # Vector + Semantic = Hybrid
                "integrated_vectorizer": bool(vectorizers),  # Required for agent VECTOR_SEMANTIC_HYBRID
                "vectorizer_deployment": AOAI_EMBED_DEPLOYMENT if has_vector_field else None
            }
            logging.info(f"Index configuration: {config_details}")
            
        except Exception as e:
            if "already exists" in str(e).lower():
                logging.info(f"Index {index_name} already exists")
            else:
                logging.error(f"Failed to create index {index_name}: {e}")
                raise
        
        return index_name


def _is_zip_file(blob_name: str) -> bool:
    """Check if a blob is a zip file."""
    return blob_name.lower().endswith(".zip")


@trace_function("extract_zip_files")
def _extract_zip_files(container: str, blob_name: str) -> List[Dict[str, str]]:
    """
    Extract files from a zip blob and return list of extracted file info.
    
    Args:
        container: The blob container name
        blob_name: The zip blob name
        
    Returns:
        List of dicts with 'filename' and 'content' keys for each extracted file
    """
    import zipfile
    
    bc = _blob_client().get_blob_client(container=container, blob=blob_name)
    data = bc.download_blob().readall()
    
    extracted_files = []
    
    try:
        f = io.BytesIO(data)
        with zipfile.ZipFile(f, 'r') as zip_ref:
            logger.info(f"  Extracting .zip file: {blob_name}")
            file_list = zip_ref.namelist()
            logger.info(f"  Found {len(file_list)} files in zip archive")
            
            for file_name in file_list:
                # Skip directories
                if file_name.endswith('/'):
                    continue
                
                # Skip hidden files and system files
                if file_name.startswith('.') or file_name.startswith('__'):
                    continue
                
                try:
                    file_data = zip_ref.read(file_name)
                    file_text = file_data.decode('utf-8', errors='ignore')
                    
                    if file_text.strip():  # Only include non-empty files
                        extracted_files.append({
                            'filename': file_name,
                            'content': file_text
                        })
                        logger.debug(f"    Extracted: {file_name} ({len(file_text)} chars)")
                    else:
                        logger.debug(f"    Skipping empty file: {file_name}")
                        
                except Exception as ex:
                    logger.warning(f"    Failed to extract {file_name}: {ex}")
                    continue
            
            logger.info(f"  Successfully extracted {len(extracted_files)} files from zip")
            
    except zipfile.BadZipFile:
        logger.error(f"  Invalid zip file: {blob_name}")
    except Exception as ex:
        logger.error(f"  Error processing zip file {blob_name}: {ex}")
    
    return extracted_files


@trace_function("download_text")
def _download_text(container: str, blob_name: str) -> str:
    bc = _blob_client().get_blob_client(container=container, blob=blob_name)
    data = bc.download_blob().readall()
    name = blob_name.lower()
    
    # Skip .zip files - they should be handled by _extract_zip_files and indexed separately
    if name.endswith(".zip"):
        logger.info(f"  Zip file detected: {blob_name} - will be processed separately")
        return ""  # Return empty - zip files are handled by _index_zip_blob
    
    # Existing file type handlers...
    if name.endswith((".txt", ".md")):
        return data.decode("utf-8", errors="ignore")
    if name.endswith(".docx"):
        from docx import Document
        f = io.BytesIO(data)
        doc = Document(f)
        return "\n".join([p.text for p in doc.paragraphs])
    if name.endswith(".pdf"):
        from pypdf import PdfReader
        f = io.BytesIO(data)
        reader = PdfReader(f)
        return "\n".join([(page.extract_text() or "") for page in reader.pages])
    
    # Enhanced Excel handling for dependency files
    if name.endswith(".xlsx"):
        # Check if this is a dependency file based on name or content
        is_dependency_file = any(keyword in name for keyword in ['dependency', 'dependencies', 'connection', 'network'])
        
        if not is_dependency_file:
            # Quick check for dependency-related headers
            from openpyxl import load_workbook
            f = io.BytesIO(data)
            wb = load_workbook(f, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(max_row=5, values_only=True):
                    row_text = ' '.join(str(cell).lower() for cell in row if cell)
                    if 'source server' in row_text or 'destination server' in row_text:
                        is_dependency_file = True
                        break
                if is_dependency_file:
                    break
        
        if is_dependency_file:
            # Parse as dependency data
            records = _parse_dependency_excel(data, container, blob_name)
            if records:
                chunks = _create_dependency_chunks(records)
                return "\n\n---DEPENDENCY_CHUNK---\n\n".join(chunks)
        
        # Default Excel handling for non-dependency files
        from openpyxl import load_workbook
        f = io.BytesIO(data)
        wb = load_workbook(f, read_only=True, data_only=True)
        lines: List[str] = []
        for ws in wb.worksheets:
            lines.append(f"# Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [(str(c) if c is not None else "") for c in row]
                line = " \t ".join(v.strip() for v in values if v is not None)
                if line.strip():
                    lines.append(line)
        return "\n".join(lines)
    
    # Fallback: treat as text
    return data.decode("utf-8", errors="ignore")


@trace_function("chunk_text")
def _chunk(text: str, size: int = 1200, overlap: int = 200) -> List[str]:
    words = text.split()
    chunks: List[str] = []
    i = 0
    while i < len(words):
        chunk_words = words[i : i + size]
        chunks.append(" ".join(chunk_words))
        i += max(1, size - overlap)
    return chunks


def _safe_doc_id(app_id: str, path: str, ci: int) -> str:
    raw = f"{app_id}|{path}|{ci}".encode("utf-8")
    return base64.urlsafe_b64encode(raw).decode("ascii")


def _parse_dependency_excel(data: bytes, container: str, blob_name: str) -> List[Dict]:
    """Parse dependency Excel file into structured records."""
    from openpyxl import load_workbook
    import hashlib
    
    f = io.BytesIO(data)
    wb = load_workbook(f, read_only=True, data_only=True)
    
    dependency_records = []
    
    for ws in wb.worksheets:
        headers = []
        header_row_found = False
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            # Look for header row containing key columns
            if not header_row_found:
                row_values = [str(cell).strip().lower() if cell else "" for cell in row]
                if any(col in ' '.join(row_values) for col in ['source server', 'destination server', 'source ip']):
                    headers = [str(cell).strip() if cell else "" for cell in row]
                    header_row_found = True
                    continue
            
            if header_row_found and any(cell for cell in row):
                # Create structured record from row data
                record = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx] and value:
                        record[headers[idx]] = str(value).strip()
                
                if record:  # Only add non-empty records
                    # Create a unique ID for this dependency record
                    record_str = json.dumps(record, sort_keys=True)
                    record_hash = hashlib.md5(record_str.encode()).hexdigest()[:8]
                    record['_id'] = f"{container}_{blob_name}_{record_hash}"
                    record['_source_file'] = f"{container}/{blob_name}"
                    dependency_records.append(record)
    
    return dependency_records

def _create_dependency_chunks(records: List[Dict]) -> List[str]:
    """Create searchable text chunks from dependency records."""
    chunks = []
    
    for record in records:
        # Create a semantic representation of the dependency
        parts = []
        
        # Build a natural language description
        if record.get('Time slot'):
            parts.append(f"Time: {record['Time slot']}")
        
        # Source information
        src_desc = []
        if record.get('Source server name'):
            src_desc.append(f"Server: {record['Source server name']}")
        if record.get('Source IP'):
            src_desc.append(f"IP: {record['Source IP']}")
        if record.get('Source application'):
            src_desc.append(f"Application: {record['Source application']}")
        if record.get('Source process'):
            src_desc.append(f"Process: {record['Source process']}")
        
        if src_desc:
            parts.append(f"SOURCE - {', '.join(src_desc)}")
        
        # Destination information
        dst_desc = []
        if record.get('Destination server name'):
            dst_desc.append(f"Server: {record['Destination server name']}")
        if record.get('Destination IP'):
            dst_desc.append(f"IP: {record['Destination IP']}")
        if record.get('Destination application'):
            dst_desc.append(f"Application: {record['Destination application']}")
        if record.get('Destination process'):
            dst_desc.append(f"Process: {record['Destination process']}")
        if record.get('Destination port'):
            dst_desc.append(f"Port: {record['Destination port']}")
        
        if dst_desc:
            parts.append(f"DESTINATION - {', '.join(dst_desc)}")
        
        # Create the chunk text
        chunk_text = " | ".join(parts)
        
        # Add structured JSON at the end for precise extraction
        chunk_text += f"\n[DEPENDENCY_DATA]{json.dumps(record, separators=(',', ':'))}[/DEPENDENCY_DATA]"
        
        chunks.append(chunk_text)
    
    # Group chunks if they're too small
    grouped_chunks = []
    current_group = []
    current_size = 0
    max_chunk_size = 600  # Smaller chunks for dependency data
    
    for chunk in chunks:
        chunk_size = len(chunk.split())
        if current_size + chunk_size > max_chunk_size and current_group:
            grouped_chunks.append("\n\n".join(current_group))
            current_group = [chunk]
            current_size = chunk_size
        else:
            current_group.append(chunk)
            current_size += chunk_size
    
    if current_group:
        grouped_chunks.append("\n\n".join(current_group))
    
    return grouped_chunks

# -----------------------------
# Public functions called by HTTP
# -----------------------------

# Batch controls: Azure Search REST cap is 1000 docs or ~16MB per request.
MAX_DOCS_PER_BATCH = int(os.getenv("SEARCH_MAX_DOCS_PER_BATCH", "500"))
# Keep some headroom below 16MB to account for HTTP and JSON overhead
MAX_BYTES_PER_BATCH = int(os.getenv("SEARCH_MAX_BYTES_PER_BATCH", str(12 * 1024 * 1024)))

def _estimate_doc_bytes(doc: Dict) -> int:
    # Estimate serialized size of a single doc
    # ensure_ascii=False keeps unicode; separators minimize overhead
    s = json.dumps(doc, ensure_ascii=False, separators=(",", ":"))
    return len(s.encode("utf-8"))

def _flush_batch(search: SearchClient, batch: List[Dict]) -> Dict[str, int]:
    if not batch:
        return {"uploaded": 0, "failed": 0}
    resp = search.upload_documents(documents=batch)
    uploaded = sum(1 for r in resp if r.succeeded)
    failed = len(resp) - uploaded
    return {"uploaded": uploaded, "failed": failed}


@trace_function("index_zip_blob")
def _index_zip_blob(app_id: str, container: str, blob_name: str, search: SearchClient) -> Dict:
    """
    Extract and index files from a zip blob separately.
    
    Each file in the zip is indexed with a path like:
    {container}/{blob_name}/{filename}
    e.g., 50000/kubernetes/kubernetes-input-50000.zip/deployment.yaml
    
    Args:
        app_id: Application ID
        container: Blob container name
        blob_name: Zip blob name
        search: SearchClient instance
        
    Returns:
        Dict with indexing results
    """
    logger.info(f"  Processing zip file for separate file indexing: {blob_name}")
    
    extracted_files = _extract_zip_files(container, blob_name)
    
    if not extracted_files:
        logger.warning(f"  No files extracted from zip: {blob_name}")
        return {"blobName": blob_name, "files": 0, "chunks": 0, "uploaded": 0, "failed": 0}
    
    # Retrieve blob metadata from the parent zip file
    zip_blob_metadata = _get_blob_metadata(container, blob_name)
    if zip_blob_metadata:
        logger.info(f"  Retrieved {len(zip_blob_metadata)} metadata key-value pairs from zip blob")
    
    total_chunks = 0
    total_uploaded = 0
    total_failed = 0
    files_processed = 0
    files_skipped_sensitive = 0
    
    batch: List[Dict] = []
    batch_bytes = 0
    
    for file_info in extracted_files:
        filename = file_info['filename']
        content = file_info['content']
        
        if not content.strip():
            continue
        
        # Create the path: {container}/{zip_blob_name}/{filename}
        # This gives us paths like: 50000/kubernetes/kubernetes-input-50000.zip/deployment.yaml
        file_path = f"{container}/{blob_name}/{filename}"
        
        # Scan for sensitive information before indexing
        should_skip, issues_found = _should_skip_document(content, file_path)
        if should_skip:
            logger.warning(f"    Extracted file {filename} contains security issues - SKIPPING")
            files_skipped_sensitive += 1
            continue
        
        files_processed += 1
        
        logger.debug(f"    Indexing extracted file: {filename} -> {file_path}")
        
        # Chunk the file content
        chunks = _chunk(content)
        total_chunks += len(chunks)
        
        # Generate embeddings for chunks
        vectors = _embed_texts(chunks)
        
        # Build metadata string combining blob metadata and extraction context
        extraction_context = {"extracted_from_zip": blob_name, "original_file": filename}
        combined_metadata = {**zip_blob_metadata, **extraction_context}
        metadata_str = _format_blob_metadata_as_string(combined_metadata)
        
        for ci, ch in enumerate(chunks):
            doc = {
                "id": _safe_doc_id(app_id, file_path, ci),
                "appId": app_id,
                "title": os.path.basename(filename),
                "content": ch,
                "source": "blob",
                "path": file_path,
                "chunkId": f"{ci}",
                # Additional metadata for better search context
                "filepath": file_path,
                "metadata": metadata_str
            }
            
            # Attach embedding if available
            vec = vectors[ci] if vectors and ci < len(vectors) else None
            if isinstance(vec, list):
                doc["contentVector"] = [float(x) for x in vec]
            
            size = _estimate_doc_bytes(doc)
            
            # Flush batch if needed
            if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
                res = _flush_batch(search, batch)
                total_uploaded += res["uploaded"]
                total_failed += res["failed"]
                batch = []
                batch_bytes = 0
            
            batch.append(doc)
            batch_bytes += size
    
    # Flush remaining batch
    if batch:
        res = _flush_batch(search, batch)
        total_uploaded += res["uploaded"]
        total_failed += res["failed"]
    
    logger.info(f"  Zip file indexing complete: {blob_name}")
    logger.info(f"    Files processed: {files_processed}")
    logger.info(f"    Files skipped (sensitive): {files_skipped_sensitive}")
    logger.info(f"    Total chunks: {total_chunks}")
    logger.info(f"    Uploaded: {total_uploaded}, Failed: {total_failed}")
    
    return {
        "blobName": blob_name,
        "files": files_processed,
        "files_skipped_sensitive": files_skipped_sensitive,
        "chunks": total_chunks,
        "uploaded": total_uploaded,
        "failed": total_failed
    }


@trace_function("index_blob")
def index_blob(app_id: str, container: str, blob_name: str) -> Dict:
    """Index a single blob with comprehensive tracing."""
    logger.info("=" * 80)
    logger.info(f"INDEXING SINGLE BLOB")
    logger.info(f"  Application ID: {app_id}")
    logger.info(f"  Container: {container}")
    logger.info(f"  Blob Name: {blob_name}")
    logger.info("=" * 80)
    
    if tracer:
        with tracer.start_as_current_span("index_blob_execution") as span:
            try:
                add_span_attributes(span, {
                    "indexer.app_id": app_id,
                    "indexer.container": container,
                    "indexer.blob_name": blob_name,
                    "indexer.operation": "single_blob",
                    "indexer.is_zip": _is_zip_file(blob_name)
                })

                # Ensure (re)created index for this app
                logger.info(f"Creating or updating search index for app: {app_id}")
                index_name = create_or_update_index(app_id)
                logger.info(f"Using search index: {index_name}")
                add_span_attributes(span, {"indexer.index_name": index_name})

                search = _get_search_client(index_name)

                # Handle zip files separately - extract and index each file
                if _is_zip_file(blob_name):
                    logger.info(f"Detected zip file - extracting and indexing files separately")
                    result = _index_zip_blob(app_id, container, blob_name, search)
                    span.set_status(Status(StatusCode.OK))
                    return result

                with tracer.start_as_current_span("download_and_process") as process_span:
                    try:
                        logger.info(f"Downloading blob content: {blob_name}")
                        text = _download_text(container, blob_name)
                        if not text or not text.strip():
                            logger.warning(f"Blob {blob_name} is empty or contains no text - skipping")
                            add_span_attributes(process_span, {"text.empty": True})
                            return {"blobName": blob_name, "chunks": 0, "uploaded": 0, "failed": 0}

                        logger.info(f"Downloaded {len(text)} characters from {blob_name}")
                        
                        add_span_attributes(process_span, {
                            "text.length": len(text),
                            "text.has_content": bool(text.strip())
                        })

                        # Scan for sensitive information before indexing
                        path = f"{container}/{blob_name}"
                        should_skip, issues_found = _should_skip_document(text, path)
                        if should_skip:
                            logger.warning(f"Blob {blob_name} contains security issues - SKIPPING")
                            add_span_attributes(process_span, {
                                "sensitive_info.detected": True,
                                "sensitive_info.count": len(issues_found),
                                "sensitive_info.skipped": True
                            })
                            process_span.add_event("blob_sensitive_skipped", {
                                "blob_name": blob_name,
                                "issues_count": len(issues_found)
                            })
                            process_span.set_status(Status(StatusCode.OK))
                            return {"blobName": blob_name, "chunks": 0, "uploaded": 0, "failed": 0, "skipped_sensitive": True}

                        if issues_found:
                            add_span_attributes(process_span, {
                                "sensitive_info.detected": True,
                                "sensitive_info.count": len(issues_found),
                                "sensitive_info.skipped": False
                            })

                        logger.info(f"Chunking text into processable segments...")
                        chunks = _chunk(text)
                        logger.info(f"Created {len(chunks)} chunks from blob content")
                        
                        logger.info(f"Generating embeddings for {len(chunks)} chunks...")
                        vectors = _embed_texts(chunks)
                        logger.info(f"Generated {sum(1 for v in vectors if v is not None)} embeddings")

                        add_span_attributes(process_span, {
                            "chunks.count": len(chunks),
                            "vectors.count": len(vectors)
                        })

                        process_span.set_status(Status(StatusCode.OK))
                        
                    except Exception as ex:
                        process_span.record_exception(ex)
                        process_span.set_status(Status(StatusCode.ERROR, str(ex)))
                        logger.error(f"Failed to download and process blob {blob_name}: {ex}", exc_info=True)
                        raise

                batch: List[Dict] = []
                batch_bytes = 0
                uploaded = 0
                failed = 0

                with tracer.start_as_current_span("upload_documents") as upload_span:
                    try:
                        logger.info(f"Preparing {len(chunks)} documents for upload to search index...")
                        
                        # Retrieve blob metadata for filtering
                        blob_metadata = _get_blob_metadata(container, blob_name)
                        metadata_str = _format_blob_metadata_as_string(blob_metadata)
                        if blob_metadata:
                            logger.info(f"Retrieved {len(blob_metadata)} metadata key-value pairs from blob")
                        
                        for ci, ch in enumerate(chunks):
                            path = f"{container}/{blob_name}"
                            doc = {
                                "id": _safe_doc_id(app_id, path, ci),
                                "appId": app_id,
                                "title": os.path.basename(blob_name),
                                "content": ch,
                                "source": "blob",
                                "path": path,
                                "chunkId": f"{ci}",
                                "filepath": path,
                                "metadata": metadata_str,
                            }
                            # attach embedding if available
                            vec = vectors[ci] if vectors and ci < len(vectors) else None
                            if isinstance(vec, list):
                                doc["contentVector"] = [float(x) for x in vec]
                            else:
                                logger.debug(f"Chunk {ci} has no vector embedding")
                            
                            size = _estimate_doc_bytes(doc)
                            if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
                                logger.info(f"Flushing batch of {len(batch)} documents to search index...")
                                res = _flush_batch(search, batch)
                                uploaded += res["uploaded"]
                                failed += res["failed"]
                                logger.info(f"  Batch upload complete: {res['uploaded']} uploaded, {res['failed']} failed")
                                batch = []
                                batch_bytes = 0
                            batch.append(doc)
                            batch_bytes += size

                        if batch:
                            logger.info(f"Flushing final batch of {len(batch)} documents to search index...")
                            res = _flush_batch(search, batch)
                            uploaded += res["uploaded"]
                            failed += res["failed"]
                            logger.info(f"  Final batch upload complete: {res['uploaded']} uploaded, {res['failed']} failed")

                        add_span_attributes(upload_span, {
                            "upload.documents_uploaded": uploaded,
                            "upload.documents_failed": failed,
                            "upload.total_chunks": len(chunks)
                        })

                        upload_span.set_status(Status(StatusCode.OK))
                        
                    except Exception as ex:
                        upload_span.record_exception(ex)
                        upload_span.set_status(Status(StatusCode.ERROR, str(ex)))
                        logger.error(f"Failed to upload documents to search index: {ex}", exc_info=True)
                        raise

                result = {"blobName": blob_name, "chunks": len(chunks), "uploaded": uploaded, "failed": failed}
                
                add_span_attributes(span, {
                    "result.chunks": len(chunks),
                    "result.uploaded": uploaded,
                    "result.failed": failed
                })

                logger.info("=" * 80)
                logger.info(f"INDEXING COMPLETE FOR BLOB: {blob_name}")
                logger.info(f"  Total Chunks: {len(chunks)}")
                logger.info(f"  Documents Uploaded: {uploaded}")
                logger.info(f"  Documents Failed: {failed}")
                logger.info(f"  Success Rate: {(uploaded / max(uploaded + failed, 1)) * 100:.1f}%")
                logger.info("=" * 80)

                span.set_status(Status(StatusCode.OK))
                return result
                
            except Exception as ex:
                span.record_exception(ex)
                span.set_status(Status(StatusCode.ERROR, str(ex)))
                logger.error(f"Failed to index blob {blob_name}: {ex}", exc_info=True)
                raise
    else:
        # Fallback without tracing - original implementation
        # Ensure (re)created index for this app
        index_name = create_or_update_index(app_id)

        search = _get_search_client(index_name)

        # Handle zip files separately - extract and index each file
        if _is_zip_file(blob_name):
            logger.info(f"Detected zip file - extracting and indexing files separately")
            return _index_zip_blob(app_id, container, blob_name, search)

        text = _download_text(container, blob_name)
        if not text or not text.strip():
            return {"blobName": blob_name, "chunks": 0, "uploaded": 0, "failed": 0}

        # Scan for sensitive information before indexing
        path = f"{container}/{blob_name}"
        should_skip, issues_found = _should_skip_document(text, path)
        if should_skip:
            logger.warning(f"Blob {blob_name} contains security issues - SKIPPING")
            return {"blobName": blob_name, "chunks": 0, "uploaded": 0, "failed": 0, "skipped_sensitive": True}

        chunks = _chunk(text)
        vectors = _embed_texts(chunks)

        batch: List[Dict] = []
        batch_bytes = 0
        uploaded = 0
        failed = 0

        for ci, ch in enumerate(chunks):
            path = f"{container}/{blob_name}"
            doc = {
                "id": _safe_doc_id(app_id, path, ci),
                "appId": app_id,
                "title": os.path.basename(blob_name),
                "content": ch,
                "source": "blob",
                "path": path,
                "chunkId": f"{ci}",
            }
            # attach embedding if available
            vec = vectors[ci] if vectors and ci < len(vectors) else None
            if isinstance(vec, list):
                doc["contentVector"] = [float(x) for x in vec]
            size = _estimate_doc_bytes(doc)
            if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
                res = _flush_batch(search, batch)
                uploaded += res["uploaded"]
                failed += res["failed"]
                batch = []
                batch_bytes = 0
            batch.append(doc)
            batch_bytes += size

        if batch:
            res = _flush_batch(search, batch)
            uploaded += res["uploaded"]
            failed += res["failed"]

        return {"blobName": blob_name, "chunks": len(chunks), "uploaded": uploaded, "failed": failed}


@trace_function("index_container")
def index_container(app_id: str, container: str, folder_prefix: str = None) -> Dict:
    """Index entire container or filtered subset with comprehensive tracing and managed identity.
    
    Args:
        app_id: Application ID for the index
        container: Container name to index from
        folder_prefix: Optional folder prefix to limit indexing to specific folder (e.g., 'kubernetes/input')
                      When folder_prefix is provided, the existing index is preserved and new documents
                      are added/updated without deleting the index. This enables incremental indexing.
    """
    # Determine if we should preserve the existing index (incremental mode)
    # When folder_prefix is provided, we only want to add/update documents, not recreate the index
    preserve_index = folder_prefix is not None
    
    logger.info("=" * 80)
    logger.info(f"INDEXING CONTAINER (with optional folder filter)")
    logger.info(f"  Application ID: {app_id}")
    logger.info(f"  Container: {container}")
    logger.info(f"  Folder Prefix: {folder_prefix}")
    logger.info(f"  Mode: {'INCREMENTAL (preserving existing index)' if preserve_index else 'FULL (recreating index)'}")
    logger.info("=" * 80)
    
    if tracer:
        with tracer.start_as_current_span("index_container_execution") as span:
            try:
                add_span_attributes(span, {
                    "indexer.app_id": app_id,
                    "indexer.container": container,
                    "indexer.operation": "incremental" if preserve_index else "full_reindex",
                    "indexer.has_folder_prefix": folder_prefix is not None,
                    "indexer.folder_prefix": folder_prefix,
                    "indexer.preserve_index": preserve_index
                })

                # Create or update index for this app
                # When folder_prefix is provided, preserve existing index (incremental mode)
                # Otherwise, force recreate the index (full reindex mode)
                if preserve_index:
                    logger.info(f"INCREMENTAL MODE: Preserving existing search index for app: {app_id}")
                    logger.info(f"  Only documents matching folder_prefix '{folder_prefix}' will be added/updated")
                else:
                    logger.info(f"FULL REINDEX MODE: Recreating search index for app: {app_id}")
                
                index_name = create_or_update_index(app_id, force_recreate=not preserve_index)
                logger.info(f"Using search index: {index_name}")
                add_span_attributes(span, {"indexer.index_name": index_name})

                search = _get_search_client(index_name)
                bs = _blob_client().get_container_client(container)

                total_uploaded = 0
                total_failed = 0
                total_chunks = 0
                total_blobs = 0

                batch: List[Dict] = []
                batch_bytes = 0

                with tracer.start_as_current_span("process_blobs") as process_span:
                    try:
                        logger.info(f"Listing blobs in container: {container}")
                        blob_list = list(bs.list_blobs())
                        logger.info(f"Found {len(blob_list)} blobs to process")
                        add_span_attributes(process_span, {"blobs.total_count": len(blob_list)})

                        for blob_idx, blob in enumerate(blob_list):
                            name = blob.name
                            
                            # Apply folder_prefix filter if specified
                            if folder_prefix and not name.startswith(folder_prefix):
                                logger.debug(f"Skipping blob {name} - does not match folder_prefix: {folder_prefix}")
                                continue
                            
                            logger.info(f"Processing blob {blob_idx + 1}/{len(blob_list)}: {name}")
                            
                            with tracer.start_as_current_span(f"process_blob_{blob_idx + 1}") as blob_span:
                                try:
                                    add_span_attributes(blob_span, {
                                        "blob.name": name,
                                        "blob.index": blob_idx + 1,
                                        "blob.size": getattr(blob, 'size', None),
                                        "blob.is_zip": _is_zip_file(name)
                                    })

                                    # Handle zip files separately - extract and index each file
                                    if _is_zip_file(name):
                                        logger.info(f"  Detected zip file - extracting and indexing files separately")
                                        zip_result = _index_zip_blob(app_id, container, name, search)
                                        total_chunks += zip_result.get("chunks", 0)
                                        total_uploaded += zip_result.get("uploaded", 0)
                                        total_failed += zip_result.get("failed", 0)
                                        total_blobs += zip_result.get("files", 0)  # Count extracted files
                                        blob_span.add_event("zip_processed", {
                                            "files": zip_result.get("files", 0),
                                            "chunks": zip_result.get("chunks", 0)
                                        })
                                        blob_span.set_status(Status(StatusCode.OK))
                                        continue

                                    logger.info(f"  Downloading blob: {name}")
                                    text = _download_text(container, name)
                                    if not text or not text.strip():
                                        logger.warning(f"  Blob {name} is empty or contains no text - skipping")
                                        blob_span.add_event("blob_empty", {"blob_name": name})
                                        continue
                                    
                                    logger.info(f"  Downloaded {len(text)} characters from {name}")
                                    
                                    # Scan for sensitive information before indexing
                                    should_skip, issues_found = _should_skip_document(text, name)
                                    if should_skip:
                                        logger.warning(f"  Blob {name} contains security issues - SKIPPING")
                                        blob_span.add_event("blob_sensitive_skipped", {
                                            "blob_name": name,
                                            "issues_count": len(issues_found)
                                        })
                                        add_span_attributes(blob_span, {
                                            "sensitive_info.detected": True,
                                            "sensitive_info.count": len(issues_found),
                                            "sensitive_info.skipped": True
                                        })
                                        continue
                                    
                                    chunks = _chunk(text)
                                    total_chunks += len(chunks)
                                    total_blobs += 1

                                    logger.info(f"  Created {len(chunks)} chunks from blob")
                                    
                                    add_span_attributes(blob_span, {
                                        "text.length": len(text),
                                        "chunks.count": len(chunks),
                                        "sensitive_info.detected": len(issues_found) > 0,
                                        "sensitive_info.count": len(issues_found)
                                    })

                                    logger.info(f"  Generating embeddings for {len(chunks)} chunks...")
                                    vectors = _embed_texts(chunks)
                                    logger.info(f"  Generated {sum(1 for v in vectors if v is not None)} embeddings")

                                    # Retrieve blob metadata for filtering
                                    blob_metadata = _get_blob_metadata(container, name)
                                    metadata_str = _format_blob_metadata_as_string(blob_metadata)
                                    if blob_metadata:
                                        logger.info(f"  Retrieved {len(blob_metadata)} metadata key-value pairs from blob")

                                    for ci, ch in enumerate(chunks):
                                        path = f"{container}/{name}"
                                        doc = {
                                            "id": _safe_doc_id(app_id, path, ci),
                                            "appId": app_id,
                                            "title": os.path.basename(name),
                                            "content": ch,
                                            "source": "blob",
                                            "path": path,
                                            "chunkId": f"{ci}",
                                            "filepath": path,
                                            "metadata": metadata_str,
                                        }
                                        vec = vectors[ci] if vectors and ci < len(vectors) else None
                                        if isinstance(vec, list):
                                            doc["contentVector"] = [float(x) for x in vec]
                                        size = _estimate_doc_bytes(doc)
                                        
                                        if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
                                            with tracer.start_as_current_span("flush_batch") as flush_span:
                                                try:
                                                    logger.info(f"  Flushing batch of {len(batch)} documents...")
                                                    res = _flush_batch(search, batch)
                                                    total_uploaded += res["uploaded"]
                                                    total_failed += res["failed"]
                                                    logger.info(f"    Uploaded: {res['uploaded']}, Failed: {res['failed']}")
                                                    
                                                    add_span_attributes(flush_span, {
                                                        "batch.documents": len(batch),
                                                        "batch.uploaded": res["uploaded"],
                                                        "batch.failed": res["failed"]
                                                    })
                                                    
                                                    flush_span.set_status(Status(StatusCode.OK))
                                                except Exception as flush_ex:
                                                    flush_span.record_exception(flush_ex)
                                                    flush_span.set_status(Status(StatusCode.ERROR, str(flush_ex)))
                                                    logger.error(f"  Failed to flush batch: {flush_ex}", exc_info=True)
                                                    raise
                                            
                                            batch = []
                                            batch_bytes = 0
                                        
                                        batch.append(doc)
                                        batch_bytes += size

                                    logger.info(f"  Completed processing blob: {name}")
                                    blob_span.set_status(Status(StatusCode.OK))
                                    
                                except Exception as blob_ex:  # pragma: no cover
                                    blob_span.record_exception(blob_ex)
                                    blob_span.set_status(Status(StatusCode.ERROR, str(blob_ex)))
                                    logger.error(f"  Failed to process blob {name}: {blob_ex}", exc_info=True)
                                    total_failed += 1

                        process_span.set_status(Status(StatusCode.OK))
                        
                    except Exception as ex:
                        process_span.record_exception(ex)
                        process_span.set_status(Status(StatusCode.ERROR, str(ex)))
                        logger.error(f"Failed to process blobs in container: {ex}", exc_info=True)
                        raise

                # Final batch flush
                if batch:
                    with tracer.start_as_current_span("final_batch_flush") as final_flush_span:
                        try:
                            logger.info(f"Flushing final batch of {len(batch)} documents...")
                            res = _flush_batch(search, batch)
                            total_uploaded += res["uploaded"]
                            total_failed += res["failed"]
                            logger.info(f"  Final batch uploaded: {res['uploaded']}, Failed: {res['failed']}")
                            
                            add_span_attributes(final_flush_span, {
                                "batch.documents": len(batch),
                                "batch.uploaded": res["uploaded"],
                                "batch.failed": res["failed"]
                            })
                            
                            final_flush_span.set_status(Status(StatusCode.OK))
                        except Exception as flush_ex:
                            final_flush_span.record_exception(flush_ex)
                            final_flush_span.set_status(Status(StatusCode.ERROR, str(flush_ex)))
                            logger.error(f"Failed to flush final batch: {flush_ex}", exc_info=True)
                            raise

                result = {"blobs": total_blobs, "chunks": total_chunks, "uploaded": total_uploaded, "failed": total_failed}
                
                add_span_attributes(span, {
                    "result.blobs": total_blobs,
                    "result.chunks": total_chunks,
                    "result.uploaded": total_uploaded,
                    "result.failed": total_failed,
                    "performance.success_rate": (total_uploaded / max(total_uploaded + total_failed, 1)) * 100
                })

                logger.info("=" * 80)
                logger.info(f"CONTAINER INDEXING COMPLETE: {container}")
                logger.info(f"  Total Blobs Processed: {total_blobs}")
                logger.info(f"  Total Chunks Created: {total_chunks}")
                logger.info(f"  Documents Uploaded: {total_uploaded}")
                logger.info(f"  Documents Failed: {total_failed}")
                logger.info(f"  Success Rate: {(total_uploaded / max(total_uploaded + total_failed, 1)) * 100:.1f}%")
                logger.info("=" * 80)

                span.set_status(Status(StatusCode.OK))
                return result
                
            except Exception as ex:
                span.record_exception(ex)
                span.set_status(Status(StatusCode.ERROR, str(ex)))
                logger.error(f"Failed to index container {container}: {ex}", exc_info=True)
                raise
    else:
        # Fallback without tracing - original implementation
        # Create or update index for this app
        # When folder_prefix is provided, preserve existing index (incremental mode)
        if preserve_index:
            logger.info(f"INCREMENTAL MODE: Preserving existing search index for app: {app_id}")
        else:
            logger.info(f"FULL REINDEX MODE: Recreating search index for app: {app_id}")
        
        index_name = create_or_update_index(app_id, force_recreate=not preserve_index)

        search = _get_search_client(index_name)
        bs = _blob_client().get_container_client(container)

        total_uploaded = 0
        total_failed = 0
        total_chunks = 0
        total_blobs = 0

        batch: List[Dict] = []
        batch_bytes = 0

        for blob in bs.list_blobs():
            name = blob.name
            
            # Apply folder_prefix filter if specified
            if folder_prefix and not name.startswith(folder_prefix):
                logger.debug(f"Skipping blob {name} - does not match folder_prefix: {folder_prefix}")
                continue
            
            try:
                # Handle zip files separately - extract and index each file
                if _is_zip_file(name):
                    logger.info(f"  Detected zip file - extracting and indexing files separately: {name}")
                    zip_result = _index_zip_blob(app_id, container, name, search)
                    total_chunks += zip_result.get("chunks", 0)
                    total_uploaded += zip_result.get("uploaded", 0)
                    total_failed += zip_result.get("failed", 0)
                    total_blobs += zip_result.get("files", 0)  # Count extracted files
                    continue

                text = _download_text(container, name)
                if not text or not text.strip():
                    continue
                
                # Scan for sensitive information before indexing
                should_skip, issues_found = _should_skip_document(text, name)
                if should_skip:
                    logger.warning(f"  Blob {name} contains security issues - SKIPPING")
                    continue
                
                chunks = _chunk(text)
                total_chunks += len(chunks)
                total_blobs += 1

                vectors = _embed_texts(chunks)
                
                # Retrieve blob metadata for filtering
                blob_metadata = _get_blob_metadata(container, name)
                metadata_str = _format_blob_metadata_as_string(blob_metadata)

                for ci, ch in enumerate(chunks):
                    path = f"{container}/{name}"
                    doc = {
                        "id": _safe_doc_id(app_id, path, ci),
                        "appId": app_id,
                        "title": os.path.basename(name),
                        "content": ch,
                        "source": "blob",
                        "path": path,
                        "chunkId": f"{ci}",
                        "filepath": path,
                        "metadata": metadata_str,
                    }
                    vec = vectors[ci] if vectors and ci < len(vectors) else None
                    if isinstance(vec, list):
                        doc["contentVector"] = [float(x) for x in vec]
                    size = _estimate_doc_bytes(doc)
                    if (len(batch) + 1 > MAX_DOCS_PER_BATCH) or (batch_bytes + size > MAX_BYTES_PER_BATCH):
                        res = _flush_batch(search, batch)
                        total_uploaded += res["uploaded"]
                        total_failed += res["failed"]
                        batch = []
                        batch_bytes = 0
                    batch.append(doc)
                    batch_bytes += size
            except Exception as e:  # pragma: no cover
                logging.exception("Failed %s: %s", name, e)
                total_failed += 1

        if batch:
            res = _flush_batch(search, batch)
            total_uploaded += res["uploaded"]
            total_failed += res["failed"]

        return {"blobs": total_blobs, "chunks": total_chunks, "uploaded": total_uploaded, "failed": total_failed}
